"""Export router - generates real CSV / Excel / PDF files."""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from auth_utils import consume_export, get_current_user
from common import fetch_for_export
from db import db
from models import ExportRequest

router = APIRouter(prefix="/export", tags=["export"])

_COLUMNS = [
    ("cin", "CIN"), ("name", "Company Name"), ("status", "Status"),
    ("company_class", "Class"), ("sector", "Sector"), ("sub_sector", "Sub-Sector"),
    ("city", "City"), ("area", "Area"), ("pin_code", "PIN"),
    ("date_of_incorporation", "Incorporated"), ("authorized_capital", "Authorized Capital"),
    ("paid_up_capital", "Paid-up Capital"), ("director_count", "Directors"),
    ("data_quality_score", "Data Quality"), ("principal_activity", "Principal Activity"),
    ("roc", "ROC"), ("address", "Registered Address"),
]


def _fmt(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return "" if value is None else value


async def _gate(user: dict):
    usage = await consume_export(user)
    if not usage["allowed"]:
        raise HTTPException(status_code=403, detail=usage.get("message", "Export not allowed on your plan."))
    return usage


@router.post("/csv")
async def export_csv(payload: ExportRequest, user: dict = Depends(get_current_user)):
    await _gate(user)
    rows = await fetch_for_export(db, payload.model_dump(), payload.limit)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([h for _, h in _COLUMNS])
    for r in rows:
        writer.writerow([_fmt(r.get(k)) for k, _ in _COLUMNS])
    buf.seek(0)
    fname = f"corpintel_companies_{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv"
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.post("/excel")
async def export_excel(payload: ExportRequest, user: dict = Depends(get_current_user)):
    await _gate(user)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    rows = await fetch_for_export(db, payload.model_dump(), payload.limit)
    wb = Workbook()
    wb.remove(wb.active)
    # group by city -> sheet per city
    by_city = {}
    for r in rows:
        by_city.setdefault(r.get("city") or "Other", []).append(r)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    if not by_city:
        by_city["Companies"] = []
    for city, crows in by_city.items():
        ws = wb.create_sheet(title=str(city)[:31])
        ws.append([h for _, h in _COLUMNS])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for r in crows:
            ws.append([_fmt(r.get(k)) for k, _ in _COLUMNS])
        for i, (_, h) in enumerate(_COLUMNS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, min(40, len(h) + 6))
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"corpintel_companies_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.post("/pdf")
async def export_pdf(payload: ExportRequest, user: dict = Depends(get_current_user)):
    await _gate(user)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)
    rows = await fetch_for_export(db, payload.model_dump(), min(payload.limit, 200))
    # summary aggregates
    by_city, by_sector = {}, {}
    for r in rows:
        by_city[r.get("city") or "Other"] = by_city.get(r.get("city") or "Other", 0) + 1
        by_sector[r.get("sector") or "Unclassified"] = by_sector.get(r.get("sector") or "Unclassified", 0) + 1
    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    elems = [Paragraph("CorpIntel India — Company Intelligence Report", styles["Title"]),
             Paragraph(datetime.now(timezone.utc).strftime("Generated %d %b %Y, %H:%M UTC"), styles["Normal"]),
             Spacer(1, 8 * mm),
             Paragraph(f"Total companies in report: <b>{len(rows)}</b>", styles["Heading2"])]
    # city breakdown table
    city_data = [["City", "Companies"]] + [[k, str(v)] for k, v in sorted(by_city.items(), key=lambda x: -x[1])]
    sector_data = [["Top Sectors", "Companies"]] + [[k, str(v)] for k, v in sorted(by_sector.items(), key=lambda x: -x[1])[:8]]
    for title, data in (("City Distribution", city_data), ("Sector Distribution", sector_data)):
        elems.append(Spacer(1, 4 * mm))
        elems.append(Paragraph(title, styles["Heading3"]))
        t = Table(data, hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elems.append(t)
    # top companies table
    elems.append(Spacer(1, 6 * mm))
    elems.append(Paragraph("Companies (top 25 by paid-up capital)", styles["Heading3"]))
    top = sorted(rows, key=lambda r: r.get("paid_up_capital", 0), reverse=True)[:25]
    comp_data = [["Name", "City", "Sector", "Paid-up (₹)"]]
    for r in top:
        comp_data.append([str(r.get("name", ""))[:34], r.get("city", ""),
                          str(r.get("sector", ""))[:20], f"{int(r.get('paid_up_capital', 0)):,}"])
    ct = Table(comp_data, colWidths=[70 * mm, 28 * mm, 40 * mm, 30 * mm])
    ct.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F4A620")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7FB")]),
    ]))
    elems.append(ct)
    doc.build(elems)
    out.seek(0)
    fname = f"corpintel_report_{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf"
    return StreamingResponse(out, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})
